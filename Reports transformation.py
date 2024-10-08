import pandas as pd
from openpyxl import load_workbook
import os.path
from shutil import copyfile
from datetime import datetime


# If the Excel for the unit does not exist, we create it
# Otherwise we expand it with the most recent monthly information
def expand_unit(value, key):
    file_path = 'By Unit/'+key+'.xlsx'
    if os.path.isfile(file_path):  # The file exists
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        pd.DataFrame(value[key]).to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row,
                                          index=False, header=False)

        writer.save()
    else:  # The file do not exist
        pd.DataFrame(value[key]).to_excel(file_path, index=False)
    return


# ##################### Processing of the monthly tracker ##################### #
print('Starting processing of the last month tracker')

#    1 - Save a copy in By Month folder
file_name = '../Monthly info needed for reports_v2_Mar24.xlsx'

copyfile(file_name,
         'By Month/Monthly info_' + str(datetime.now().year) + '_' + str(datetime.now().month) + '.xlsx')

#    2 - Read the file
xlsx = pd.ExcelFile(file_name)
sheets = {'World Bank': pd.read_excel(xlsx, 'WB'),
          'DRR': pd.read_excel(xlsx, 'DRR'),
          'GFA_FFC_RR': pd.read_excel(xlsx, 'GFA'),
          'DS&I': pd.read_excel(xlsx, 'DSI'),
          'School Feeding': pd.read_excel(xlsx, 'School Feeding'),
          'Nutrition': pd.read_excel(xlsx, 'Nutrition'),
          'Self-Reliance': pd.read_excel(xlsx, 'Self-Reliance'),
          'Livelihoods': pd.read_excel(xlsx, 'Livelihoods'),
          'FSS': pd.read_excel(xlsx, 'FSS'),
          'ETS': pd.read_excel(xlsx, 'TEC Common Services'),
          'Supply Chain': pd.read_excel(xlsx, 'Supply Chain'),
          'EPR': pd.read_excel(xlsx, 'EPR'),
          'Protection, Gender and DI': pd.read_excel(xlsx, 'Protection, Gender and DI'),
          'Eng': pd.read_excel(xlsx, 'Eng'),
          'PRC and MEAL': pd.read_excel(xlsx, 'RAM-PRC')}

#    3 - Fill the general database
if not os.path.isfile("General Database.xlsx",):  # New base
    pd.concat(sheets.values(), ignore_index=True).to_excel("General Database.xlsx",
                                                           sheet_name='Sheet1', index=False)
else:  # Update base
    book2 = load_workbook('General Database.xlsx')
    writ = pd.ExcelWriter('General Database.xlsx', engine='openpyxl')
    writ.book = book2
    writ.sheets = {ws.title: ws for ws in book2.worksheets}

    pd.concat(sheets.values(), ignore_index=True).to_excel(writ, sheet_name='Sheet1',
                                                           startrow=writ.sheets['Sheet1'].max_row,
                                                           index=False, header=False)
    writ.save()

#    4 - Fill each unit database
for sheet in sheets:
    print('Processing information for ' + sheet)
    expand_unit(sheets, sheet)

print('Data transformation complete')  # End
